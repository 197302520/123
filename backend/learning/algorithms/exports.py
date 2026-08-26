from __future__ import annotations

import base64
import csv
import io
import json
import xml.etree.ElementTree as ET
from typing import Any

from .errors import AlgorithmInputError
from .graph import normalize_graph


FORMATS = {"json", "csv", "xlsx", "graphml", "gexf", "gml", "pajek", "edgelist", "adjacency"}
XLSX_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
GRAPHML_SCHEMA_MARKER = "sna_graphspec_v1"
GRAPHML_ATTRIBUTES_KEY = "sna_attributes_json"


def _sorted_graph(graph: dict[str, Any]) -> dict[str, Any]:
    return {
        "directed": graph["directed"],
        "nodes": sorted(graph["nodes"], key=lambda node: node["id"]),
        "edges": sorted(graph["edges"], key=lambda edge: (edge["source"], edge["target"], edge["weight"])),
    }


def _csv_edges(graph: dict[str, Any]) -> str:
    output = io.StringIO(newline="")
    writer = csv.writer(output, lineterminator="\n")
    writer.writerow(["source", "target", "weight"])
    for edge in graph["edges"]:
        writer.writerow([edge["source"], edge["target"], edge["weight"]])
    return output.getvalue()


def _adjacency(graph: dict[str, Any]) -> str:
    nodes = [node["id"] for node in graph["nodes"]]
    weights = {(edge["source"], edge["target"]): edge["weight"] for edge in graph["edges"]}
    if not graph["directed"]:
        weights.update({(target, source): value for (source, target), value in list(weights.items())})
    output = io.StringIO(newline="")
    writer = csv.writer(output, lineterminator="\n")
    writer.writerow(["node", *nodes])
    for source in nodes:
        writer.writerow([source, *[weights.get((source, target), 0.0) for target in nodes]])
    return output.getvalue()


def _graphml(graph: dict[str, Any]) -> str:
    namespace = "http://graphml.graphdrawing.org/xmlns"
    ET.register_namespace("", namespace)
    root = ET.Element(f"{{{namespace}}}graphml")
    ET.SubElement(root, f"{{{namespace}}}key", {"id": "label", "for": "node", "attr.name": "label", "attr.type": "string"})
    ET.SubElement(root, f"{{{namespace}}}key", {"id": GRAPHML_ATTRIBUTES_KEY, "for": "node", "attr.name": GRAPHML_ATTRIBUTES_KEY, "attr.type": "string"})
    ET.SubElement(root, f"{{{namespace}}}key", {"id": "sna_schema", "for": "graph", "attr.name": "sna_schema", "attr.type": "string"})
    ET.SubElement(root, f"{{{namespace}}}key", {"id": "weight", "for": "edge", "attr.name": "weight", "attr.type": "double"})
    graph_element = ET.SubElement(root, f"{{{namespace}}}graph", {"id": "G", "edgedefault": "directed" if graph["directed"] else "undirected"})
    ET.SubElement(graph_element, f"{{{namespace}}}data", {"key": "sna_schema"}).text = GRAPHML_SCHEMA_MARKER
    for node in graph["nodes"]:
        element = ET.SubElement(graph_element, f"{{{namespace}}}node", {"id": node["id"]})
        ET.SubElement(element, f"{{{namespace}}}data", {"key": "label"}).text = node["label"]
        if node.get("attributes"):
            ET.SubElement(element, f"{{{namespace}}}data", {"key": GRAPHML_ATTRIBUTES_KEY}).text = json.dumps(
                node["attributes"], ensure_ascii=False, sort_keys=True,
            )
    for index, edge in enumerate(graph["edges"]):
        element = ET.SubElement(graph_element, f"{{{namespace}}}edge", {"id": f"e{index}", "source": edge["source"], "target": edge["target"]})
        ET.SubElement(element, f"{{{namespace}}}data", {"key": "weight"}).text = format(edge["weight"], ".17g")
    return ET.tostring(root, encoding="unicode", xml_declaration=True)


def _gexf(graph: dict[str, Any]) -> str:
    namespace = "http://gexf.net/1.3"
    ET.register_namespace("", namespace)
    root = ET.Element(f"{{{namespace}}}gexf", {"version": "1.3"})
    graph_element = ET.SubElement(root, f"{{{namespace}}}graph", {"mode": "static", "defaultedgetype": "directed" if graph["directed"] else "undirected"})
    nodes_element = ET.SubElement(graph_element, f"{{{namespace}}}nodes")
    for node in graph["nodes"]:
        ET.SubElement(nodes_element, f"{{{namespace}}}node", {"id": node["id"], "label": node["label"]})
    edges_element = ET.SubElement(graph_element, f"{{{namespace}}}edges")
    for index, edge in enumerate(graph["edges"]):
        ET.SubElement(edges_element, f"{{{namespace}}}edge", {"id": str(index), "source": edge["source"], "target": edge["target"], "weight": format(edge["weight"], ".17g")})
    return ET.tostring(root, encoding="unicode", xml_declaration=True)


def _quote(value: str) -> str:
    return '"' + value.replace("\\", "\\\\").replace('"', '\\"') + '"'


def _gml(graph: dict[str, Any]) -> str:
    index = {node["id"]: offset for offset, node in enumerate(graph["nodes"])}
    lines = ["graph [", f"  directed {1 if graph['directed'] else 0}"]
    for node in graph["nodes"]:
        lines.extend(["  node [", f"    id {index[node['id']]}", f"    label {_quote(node['label'])}", "  ]"])
    for edge in graph["edges"]:
        lines.extend(["  edge [", f"    source {index[edge['source']]}", f"    target {index[edge['target']]}", f"    weight {format(edge['weight'], '.17g')}", "  ]"])
    lines.append("]")
    return "\n".join(lines) + "\n"


def _pajek(graph: dict[str, Any]) -> str:
    index = {node["id"]: offset + 1 for offset, node in enumerate(graph["nodes"])}
    lines = [f"*Vertices {len(graph['nodes'])}"]
    lines.extend(f"{index[node['id']]} {_quote(node['label'])}" for node in graph["nodes"])
    lines.append("*Arcs" if graph["directed"] else "*Edges")
    lines.extend(f"{index[edge['source']]} {index[edge['target']]} {format(edge['weight'], '.17g')}" for edge in graph["edges"])
    return "\n".join(lines) + "\n"


def _xlsx_bytes(graph: dict[str, Any]) -> bytes:
    """Native Excel workbook: node list, weighted adjacency matrix, and the edge table."""
    from openpyxl import Workbook

    nodes = [node["id"] for node in graph["nodes"]]
    weights = {(edge["source"], edge["target"]): edge["weight"] for edge in graph["edges"]}
    if not graph["directed"]:
        weights.update({(target, source): value for (source, target), value in list(weights.items())})
    workbook = Workbook()
    node_sheet = workbook.active
    node_sheet.title = "节点编号清单"
    node_sheet.append(["编号", "标签"])
    for node in graph["nodes"]:
        node_sheet.append([node["id"], node["label"]])
    matrix_sheet = workbook.create_sheet("邻接矩阵")
    matrix_sheet.append(["node", *nodes])
    for source in nodes:
        matrix_sheet.append([source, *[weights.get((source, target), 0.0) for target in nodes]])
    edge_sheet = workbook.create_sheet("边列表")
    edge_sheet.append(["source", "target", "weight"])
    for edge in graph["edges"]:
        edge_sheet.append([edge["source"], edge["target"], edge["weight"]])
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def export_graph(graph: dict[str, Any], format_name: str) -> dict[str, str]:
    normalized = _sorted_graph(normalize_graph(graph))
    format_name = format_name.lower()
    if format_name not in FORMATS:
        raise AlgorithmInputError(f"不支持的导出格式：{format_name}。", path="parameters.format")
    if format_name == "json":
        content = json.dumps(normalized, ensure_ascii=False, sort_keys=True, indent=2)
        mime_type = "application/json"
        extension = "json"
        encoding = "text"
    elif format_name == "csv":
        content, mime_type, extension, encoding = _csv_edges(normalized), "text/csv", "csv", "text"
    elif format_name == "adjacency":
        content, mime_type, extension, encoding = _adjacency(normalized), "text/csv", "adjacency.csv", "text"
    elif format_name == "xlsx":
        content, mime_type, extension, encoding = (
            base64.b64encode(_xlsx_bytes(normalized)).decode("ascii"), XLSX_MIME_TYPE, "xlsx", "base64",
        )
    elif format_name == "graphml":
        content, mime_type, extension, encoding = _graphml(normalized), "application/graphml+xml", "graphml", "text"
    elif format_name == "gexf":
        content, mime_type, extension, encoding = _gexf(normalized), "application/gexf+xml", "gexf", "text"
    elif format_name == "gml":
        content, mime_type, extension, encoding = _gml(normalized), "text/plain", "gml", "text"
    elif format_name == "pajek":
        content, mime_type, extension, encoding = _pajek(normalized), "text/plain", "net", "text"
    else:
        content = "\n".join(f"{edge['source']} {edge['target']} {format(edge['weight'], '.17g')}" for edge in normalized["edges"]) + ("\n" if normalized["edges"] else "")
        mime_type, extension, encoding = "text/plain", "edgelist", "text"
    return {"format": format_name, "mime_type": mime_type, "filename": f"network.{extension}", "encoding": encoding, "content": content}
