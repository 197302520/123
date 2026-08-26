from __future__ import annotations

import csv
import io
import json
import re
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Any, Iterable, Sequence

import networkx as nx
from django.conf import settings

from .algorithms.errors import AlgorithmInputError
from .algorithms.graph import nx_to_graph, normalize_graph


ALLOWED_SUFFIXES = {".txt", ".csv", ".xlsx", ".json", ".graphml", ".gexf"}
ARCHIVE_MAGIC = (b"PK\x03\x04", b"PK\x05\x06", b"PK\x07\x08")


class UnsafeUploadError(ValueError):
    def __init__(self, message: str, *, status_code: int = 400):
        super().__init__(message)
        self.status_code = status_code


def _decode(data: bytes) -> str:
    if b"\x00" in data:
        raise UnsafeUploadError("文件包含不支持的二进制内容。")
    try:
        return data.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise UnsafeUploadError("文本文件必须使用 UTF-8 编码。") from exc


def _edge_graph(rows: Iterable[Sequence[Any]]) -> dict[str, Any]:
    edges: list[dict[str, Any]] = []
    nodes: set[str] = set()
    max_edges = int(getattr(settings, "PUBLIC_MAX_EDGES", 20_000))
    max_nodes = int(getattr(settings, "PUBLIC_MAX_NODES", 2_000))
    for index, row in enumerate(rows):
        if not row or all(str(value).strip() == "" for value in row):
            continue
        if index == 0 and len(row) >= 2 and str(row[0]).strip().lower() == "source" and str(row[1]).strip().lower() == "target":
            continue
        if len(row) < 2:
            raise UnsafeUploadError(f"第 {index + 1} 行至少需要 source 和 target。")
        if len(edges) >= max_edges:
            raise UnsafeUploadError("导入图超过公开边数限制。", status_code=413)
        source, target = str(row[0]).strip(), str(row[1]).strip()
        if not source or not target:
            raise UnsafeUploadError(f"第 {index + 1} 行的节点不能为空。")
        raw_weight = row[2] if len(row) > 2 and str(row[2]).strip() else 1
        try:
            weight = float(raw_weight)
        except (TypeError, ValueError) as exc:
            raise UnsafeUploadError(f"第 {index + 1} 行的权重必须是数值。") from exc
        nodes.update((source, target))
        if len(nodes) > max_nodes:
            raise UnsafeUploadError("导入图超过公开节点数限制。", status_code=413)
        edges.append({"source": source, "target": target, "weight": weight})
    return normalize_graph({
        "directed": False,
        "nodes": [{"id": node} for node in sorted(nodes)],
        "edges": edges,
    })


def _parse_xlsx(data: bytes) -> dict[str, Any]:
    if not data.startswith(ARCHIVE_MAGIC):
        raise UnsafeUploadError("XLSX 文件结构无效。")
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            infos = archive.infolist()
            names = [info.filename.replace("\\", "/").lower() for info in infos]
            expanded_limit = int(getattr(settings, "MAX_UPLOAD_BYTES", 20 * 1024 * 1024)) * 4
            if len(infos) > 2_000 or sum(info.file_size for info in infos) > expanded_limit:
                raise UnsafeUploadError("XLSX 解压后内容超过安全限制。", status_code=413)
            if any(
                "vbaproject.bin" in name or name.startswith("xl/externallinks/")
                or name.startswith("/") or ".." in name.split("/")
                for name in names
            ):
                raise UnsafeUploadError("XLSX 不得包含宏、外部链接或越界路径。", status_code=415)
        from openpyxl import load_workbook
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True, keep_links=False)
        try:
            parsed = _edge_graph(workbook.active.iter_rows(min_col=1, max_col=3, values_only=True))
        finally:
            workbook.close()
    except UnsafeUploadError:
        raise
    except (ImportError, KeyError, OSError, ValueError, zipfile.BadZipFile) as exc:
        raise UnsafeUploadError("无法安全解析 XLSX 文件。") from exc
    return parsed


def _skip_json_space(text: str, position: int) -> int:
    while position < len(text) and text[position] in " \t\r\n":
        position += 1
    return position


def _parse_bounded_json_graph(text: str) -> dict[str, Any]:
    """Decode graph arrays one item at a time so public caps bound retained objects."""
    decoder = json.JSONDecoder()
    position = _skip_json_space(text, 0)
    if position >= len(text) or text[position] != "{":
        raise UnsafeUploadError("JSON 图必须是对象。")
    position += 1
    payload: dict[str, Any] = {}
    allowed = {"directed", "nodes", "edges"}
    while True:
        position = _skip_json_space(text, position)
        if position < len(text) and text[position] == "}":
            position += 1
            break
        key, position = decoder.raw_decode(text, position)
        if not isinstance(key, str) or key not in allowed or key in payload:
            raise UnsafeUploadError("JSON 图包含重复或不支持的字段。")
        position = _skip_json_space(text, position)
        if position >= len(text) or text[position] != ":":
            raise UnsafeUploadError("JSON 图字段缺少冒号。")
        position = _skip_json_space(text, position + 1)
        if key in {"nodes", "edges"}:
            if position >= len(text) or text[position] != "[":
                raise UnsafeUploadError(f"{key} 必须是数组。")
            position += 1
            values: list[Any] = []
            limit = int(getattr(settings, "PUBLIC_MAX_NODES" if key == "nodes" else "PUBLIC_MAX_EDGES", 2_000 if key == "nodes" else 20_000))
            while True:
                position = _skip_json_space(text, position)
                if position < len(text) and text[position] == "]":
                    position += 1
                    break
                value, position = decoder.raw_decode(text, position)
                if len(values) >= limit:
                    raise UnsafeUploadError(f"导入图超过公开{'节点' if key == 'nodes' else '边'}数限制。", status_code=413)
                values.append(value)
                position = _skip_json_space(text, position)
                if position < len(text) and text[position] == ",":
                    position += 1
                    continue
                if position < len(text) and text[position] == "]":
                    position += 1
                    break
                raise UnsafeUploadError(f"{key} 数组结构无效。")
            payload[key] = values
        else:
            payload[key], position = decoder.raw_decode(text, position)
        position = _skip_json_space(text, position)
        if position < len(text) and text[position] == ",":
            position += 1
            continue
        if position < len(text) and text[position] == "}":
            position += 1
            break
        raise UnsafeUploadError("JSON 图对象结构无效。")
    if _skip_json_space(text, position) != len(text):
        raise UnsafeUploadError("JSON 图后包含额外内容。")
    return normalize_graph(payload)


def _enforce_xml_graph_limits(data: bytes) -> None:
    maximums = {
        "node": int(getattr(settings, "PUBLIC_MAX_NODES", 2_000)),
        "edge": int(getattr(settings, "PUBLIC_MAX_EDGES", 20_000)),
    }
    counts = {"node": 0, "edge": 0}
    for _, element in ET.iterparse(io.BytesIO(data), events=("end",)):
        local = element.tag.rsplit("}", 1)[-1]
        if local in counts:
            counts[local] += 1
            if counts[local] > maximums[local]:
                raise UnsafeUploadError(
                    f"导入图超过公开{'节点' if local == 'node' else '边'}数限制。", status_code=413,
                )
        element.clear()


def _parse_xml_graph(data: bytes, suffix: str) -> dict[str, Any]:
    upper = data.upper()
    if b"<!DOCTYPE" in upper or b"<!ENTITY" in upper:
        raise UnsafeUploadError("XML 图文件不得包含 DTD 或外部实体。")
    try:
        _enforce_xml_graph_limits(data)
        if suffix == ".graphml":
            network = nx.read_graphml(io.BytesIO(data))
            if network.is_multigraph():
                raise UnsafeUploadError("暂不支持多重边图。")
            return normalize_graph(nx_to_graph(network))
        root = ET.fromstring(data)
        local = lambda element: element.tag.rsplit("}", 1)[-1]
        graph_element = next((element for element in root.iter() if local(element) == "graph"), None)
        if graph_element is None:
            raise UnsafeUploadError("GEXF 缺少 graph 元素。")
        node_elements = [element for element in graph_element.iter() if local(element) == "node"]
        edge_elements = [element for element in graph_element.iter() if local(element) == "edge"]
        graph = {
            "directed": graph_element.attrib.get("defaultedgetype", "undirected") == "directed",
            "nodes": [
                {"id": element.attrib.get("id"), "label": element.attrib.get("label", element.attrib.get("id"))}
                for element in node_elements
            ],
            "edges": [
                {
                    "source": element.attrib.get("source"), "target": element.attrib.get("target"),
                    "weight": float(element.attrib.get("weight", 1)),
                }
                for element in edge_elements
            ],
        }
        return normalize_graph(graph)
    except UnsafeUploadError:
        raise
    except Exception as exc:
        raise UnsafeUploadError("无法安全解析 XML 图文件。") from exc


def parse_uploaded_graph(uploaded) -> dict[str, Any]:
    suffix = Path(uploaded.name or "").suffix.lower()
    if suffix not in ALLOWED_SUFFIXES:
        raise UnsafeUploadError("仅支持 TXT、CSV、XLSX、JSON、GraphML 和 GEXF。", status_code=415)
    limit = int(getattr(settings, "MAX_UPLOAD_BYTES", 20 * 1024 * 1024))
    if uploaded.size > limit:
        raise UnsafeUploadError("上传文件超过 20 MB 限制。", status_code=413)
    data = uploaded.read(limit + 1)
    if len(data) > limit:
        raise UnsafeUploadError("上传文件超过 20 MB 限制。", status_code=413)
    if data.startswith(ARCHIVE_MAGIC) and suffix != ".xlsx":
        raise UnsafeUploadError("不接受压缩包或伪装的归档文件。", status_code=415)
    try:
        if suffix == ".json":
            return _parse_bounded_json_graph(_decode(data))
        if suffix == ".csv":
            return _edge_graph(csv.reader(io.StringIO(_decode(data))))
        if suffix == ".txt":
            rows = (re.split(r"[\s,]+", line.strip()) for line in _decode(data).splitlines())
            return _edge_graph(rows)
        if suffix == ".xlsx":
            return _parse_xlsx(data)
        return _parse_xml_graph(data, suffix)
    except (json.JSONDecodeError, AlgorithmInputError) as exc:
        raise UnsafeUploadError(str(exc)) from exc
